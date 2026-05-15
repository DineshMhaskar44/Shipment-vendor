"""Excel & PDF helpers — bulk import, exports, and PDF report generation.

We use openpyxl directly (no pandas dependency at runtime) for tighter control
over headers/dates. PDFs use reportlab.
"""
from io import BytesIO
from datetime import datetime, date
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
)


# Column order shipped in exports / expected in imports.
SHIPMENT_COLUMNS = [
    ("bank_name", "Bank Name"),
    ("customer_po_date", "Customer PO Date"),
    ("customer_po_number", "Customer PO Number"),
    ("oem_po_date", "OEM PO Date"),
    ("oem_po_number", "OEM PO Number"),
    ("advance_payment_date", "Advance Payment Date"),
    ("balance_payment_date", "Balance Payment Date"),
    ("payment_days", "Payment Days"),
    ("delay_status", "Status (Delayed/Non Delayed)"),
    ("quantity", "Quantity"),
    ("payment_status", "Payment Status"),
    ("oem_name", "OEM Name"),
    ("device_model", "Device Model"),
    ("boe_number", "BOE Number"),
    ("shipment_quantity", "Shipment Quantity"),
    ("branding_specification", "Branding Specification"),
    ("oem_readiness_date", "OEM Readiness Date"),
    ("revised_oem_readiness_date", "Revised OEM Readiness Date"),
    ("readiness_delay_days", "Readiness Delay Days"),
    ("warehouse_gatein_date", "Warehouse Gate-in Date"),
    ("readiness_to_warehouse_days", "Readiness To Warehouse Days"),
    ("logistics_partner_name", "Logistics Partner Name"),
    ("committed_sailing_date", "Committed Sailing Date"),
    ("actual_sailing_date", "Actual Sailing Date"),
    ("sailing_delayed_days", "Sailing Delayed Days"),
    ("shipment_mode_confirmation", "Shipment Mode Confirmation"),
    ("mode", "Mode"),
    ("india_port_landing_date", "India Port Landing Date"),
    ("custom_clearance_date", "Custom Clearance Date"),
    ("cc_delayed_days", "CC Delayed Days"),
    ("proposed_eta_warehouse", "Proposed ETA at Warehouse"),
    ("actual_delivery_date", "Actual Delivery Date"),
    ("total_clearance_days", "Total Clearance Days"),
    ("dispatch_date", "Dispatch Date"),
    ("status", "Status"),
    ("handover_remarks", "Handover Remarks"),
]

DATE_FIELDS = {k for k, _ in SHIPMENT_COLUMNS if k.endswith("_date")}
INT_FIELDS = {
    "quantity", "shipment_quantity", "payment_days",
    "readiness_delay_days", "readiness_to_warehouse_days",
    "sailing_delayed_days", "cc_delayed_days", "total_clearance_days",
}


# --------------------------------------------------------------------------- #
#  Excel export
# --------------------------------------------------------------------------- #
def _autosize(ws):
    for column_cells in ws.columns:
        length = max(len(str(c.value)) if c.value is not None else 0
                     for c in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = min(length + 2, 40)


def shipments_to_xlsx(shipments) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Shipments"

    headers = [label for _, label in SHIPMENT_COLUMNS]
    ws.append(headers)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="0D6EFD")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for s in shipments:
        row = []
        for attr, _ in SHIPMENT_COLUMNS:
            val = getattr(s, attr, None)
            if isinstance(val, (datetime, date)):
                val = val.strftime("%Y-%m-%d")
            row.append(val)
        ws.append(row)

    _autosize(ws)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def vendors_to_xlsx(vendors) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Vendors"
    ws.append(["ID", "Company", "Contact", "Email", "Phone",
               "Category", "Approved", "Rating", "Created"])
    for v in vendors:
        ws.append([
            v.id, v.company_name, v.contact_person, v.email, v.phone,
            v.category, "Yes" if v.is_approved else "No", v.rating,
            v.created_at.strftime("%Y-%m-%d") if v.created_at else "",
        ])
    _autosize(ws)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def quotations_compare_xlsx(rfq, quotations) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = f"RFQ {rfq.rfq_number}"
    ws.append([f"RFQ Comparison — {rfq.rfq_number}: {rfq.title}"])
    ws.append([])
    ws.append(["Vendor", "Unit Price", "Total Price", "Currency",
               "Delivery (days)", "Payment Terms", "Warranty",
               "Submitted", "Selected"])
    for q in quotations:
        ws.append([
            q.vendor.company_name,
            float(q.unit_price) if q.unit_price else None,
            float(q.total_price) if q.total_price else None,
            q.currency,
            q.delivery_days,
            q.payment_terms,
            q.warranty,
            q.submitted_at.strftime("%Y-%m-%d %H:%M") if q.submitted_at else "",
            "YES" if q.is_selected else "",
        ])
    _autosize(ws)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# --------------------------------------------------------------------------- #
#  Excel import
# --------------------------------------------------------------------------- #
def parse_shipment_xlsx(file_storage):
    """Yield dicts (one per row) ready to feed into a Shipment(...).

    Validates the header row matches our expected column labels (lenient on
    case + whitespace).
    """
    wb = load_workbook(file_storage, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return

    header_row = rows[0]
    label_to_attr = {label.strip().lower(): attr for attr, label in SHIPMENT_COLUMNS}

    col_map = {}
    for idx, header in enumerate(header_row):
        if header is None:
            continue
        key = str(header).strip().lower()
        if key in label_to_attr:
            col_map[idx] = label_to_attr[key]

    for row in rows[1:]:
        if not any(c not in (None, "") for c in row):
            continue
        record = {}
        for idx, attr in col_map.items():
            value = row[idx] if idx < len(row) else None
            if value in ("", None):
                continue
            if attr in DATE_FIELDS:
                if isinstance(value, datetime):
                    value = value.date()
                elif isinstance(value, date):
                    pass
                else:
                    try:
                        value = datetime.strptime(str(value)[:10], "%Y-%m-%d").date()
                    except ValueError:
                        continue
            elif attr in INT_FIELDS:
                try:
                    value = int(value)
                except (ValueError, TypeError):
                    continue
            record[attr] = value
        yield record


# --------------------------------------------------------------------------- #
#  PDF reports
# --------------------------------------------------------------------------- #
def _pdf_table(headers, rows):
    data = [headers] + rows
    table = Table(data, repeatRows=1)
    style = TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0D6EFD")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1),
         [colors.whitesmoke, colors.white]),
    ])
    table.setStyle(style)
    return table


def shipments_to_pdf(shipments, title="Shipment Report") -> BytesIO:
    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=20, rightMargin=20,
                            topMargin=24, bottomMargin=24)
    styles = getSampleStyleSheet()
    story = [Paragraph(title, styles["Title"]),
             Paragraph(datetime.utcnow().strftime("Generated %Y-%m-%d %H:%M UTC"),
                       styles["Normal"]),
             Spacer(1, 12)]

    headers = ["PO #", "OEM", "Model", "Qty", "Mode",
               "Sailed", "ETA", "Delivered", "Status", "Delay"]
    rows = []
    for s in shipments:
        rows.append([
            s.customer_po_number or "",
            s.oem_name or "",
            s.device_model or "",
            s.shipment_quantity or s.quantity or "",
            s.mode or "",
            s.actual_sailing_date.strftime("%Y-%m-%d") if s.actual_sailing_date else "",
            s.proposed_eta_warehouse.strftime("%Y-%m-%d") if s.proposed_eta_warehouse else "",
            s.actual_delivery_date.strftime("%Y-%m-%d") if s.actual_delivery_date else "",
            s.status or "",
            s.delay_status or "",
        ])

    story.append(_pdf_table(headers, rows))
    doc.build(story)
    buf.seek(0)
    return buf


def quotations_compare_pdf(rfq, quotations) -> BytesIO:
    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=24, rightMargin=24,
                            topMargin=24, bottomMargin=24)
    styles = getSampleStyleSheet()
    story = [
        Paragraph(f"RFQ {rfq.rfq_number} — Vendor Comparison", styles["Title"]),
        Paragraph(rfq.title, styles["Heading3"]),
        Spacer(1, 8),
        Paragraph(rfq.product_details or "", styles["Normal"]),
        Spacer(1, 12),
    ]
    headers = ["Vendor", "Unit Price", "Total", "Curr",
               "Days", "Terms", "Warranty", "Selected"]
    rows = []
    for q in quotations:
        rows.append([
            q.vendor.company_name,
            f"{q.unit_price:.2f}" if q.unit_price else "",
            f"{q.total_price:.2f}" if q.total_price else "",
            q.currency or "",
            q.delivery_days or "",
            (q.payment_terms or "")[:30],
            (q.warranty or "")[:20],
            "YES" if q.is_selected else "",
        ])
    story.append(_pdf_table(headers, rows))
    doc.build(story)
    buf.seek(0)
    return buf
