"""Reports blueprint — six MIS reports + Excel exports.

Reports:
  1. Shipment delay analysis
  2. Vendor performance (logistics partners)
  3. OEM performance
  4. Payment pending
  5. Logistics delays
  6. Monthly MIS
"""
from datetime import datetime, date, timedelta
from io import BytesIO

from flask import (Blueprint, render_template, request, send_file,
                   url_for, redirect)
from flask_login import login_required
from sqlalchemy import func, case, extract

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from ..extensions import db
from ..models import Shipment, Quotation, RFQ, Vendor
from ..utils.decorators import staff_or_admin_required

reports_bp = Blueprint("reports", __name__,
                       template_folder="../templates/reports")


def _wb_from_rows(title, headers, rows):
    wb = Workbook()
    ws = wb.active
    ws.title = title[:30]
    ws.append([title])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])
    ws.append(headers)
    for cell in ws[3]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="0D6EFD")
        cell.alignment = Alignment(horizontal="center")
    for r in rows:
        ws.append(r)
    for col in ws.columns:
        m = max(len(str(c.value)) if c.value is not None else 0 for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(m + 2, 40)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# --------------------------------------------------------------------------- #
@reports_bp.route("/")
@login_required
@staff_or_admin_required
def index():
    return render_template("reports/index.html")


# 1. Shipment delay analysis
def _delay_rows():
    return (db.session.query(
        Shipment.id, Shipment.customer_po_number, Shipment.oem_name,
        Shipment.logistics_partner_name,
        Shipment.readiness_delay_days, Shipment.sailing_delayed_days,
        Shipment.cc_delayed_days, Shipment.delay_status, Shipment.status,
    ).filter(Shipment.delay_status == "Delayed").all())


@reports_bp.route("/delay-analysis")
@login_required
@staff_or_admin_required
def delay_analysis():
    rows = _delay_rows()
    return render_template("reports/delay_analysis.html", rows=rows)


@reports_bp.route("/delay-analysis.xlsx")
@login_required
@staff_or_admin_required
def delay_analysis_xlsx():
    rows = _delay_rows()
    headers = ["ID", "PO #", "OEM", "Logistics", "Readiness Delay",
               "Sailing Delay", "CC Delay", "Delay Status", "Status"]
    buf = _wb_from_rows("Delay Analysis", headers,
                        [list(r) for r in rows])
    return send_file(buf, as_attachment=True,
                     download_name="delay_analysis.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# 2. Vendor performance
def _vendor_perf_rows():
    return (db.session.query(
        Shipment.logistics_partner_name,
        func.count(Shipment.id).label("total"),
        func.sum(case((Shipment.delay_status == "Delayed", 1), else_=0))
            .label("delayed"),
        func.avg(Shipment.sailing_delayed_days).label("avg_sail_delay"),
        func.avg(Shipment.cc_delayed_days).label("avg_cc_delay"),
    ).filter(Shipment.logistics_partner_name.isnot(None))
     .group_by(Shipment.logistics_partner_name)
     .order_by(func.count(Shipment.id).desc()).all())


@reports_bp.route("/vendor-performance")
@login_required
@staff_or_admin_required
def vendor_performance():
    rows = _vendor_perf_rows()
    return render_template("reports/vendor_performance.html", rows=rows)


@reports_bp.route("/vendor-performance.xlsx")
@login_required
@staff_or_admin_required
def vendor_performance_xlsx():
    rows = _vendor_perf_rows()
    headers = ["Vendor", "Total Shipments", "Delayed",
               "Avg Sailing Delay (days)", "Avg CC Delay (days)"]
    rows_out = [[r[0], int(r[1]), int(r[2] or 0),
                 round(float(r[3] or 0), 1), round(float(r[4] or 0), 1)]
                for r in rows]
    buf = _wb_from_rows("Vendor Performance", headers, rows_out)
    return send_file(buf, as_attachment=True,
                     download_name="vendor_performance.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# 3. OEM performance
def _oem_perf_rows():
    return (db.session.query(
        Shipment.oem_name,
        func.count(Shipment.id).label("total"),
        func.sum(case((Shipment.status == "Delivered", 1), else_=0))
            .label("delivered"),
        func.sum(case((Shipment.delay_status == "Delayed", 1), else_=0))
            .label("delayed"),
        func.avg(Shipment.readiness_delay_days).label("avg_readiness_delay"),
    ).filter(Shipment.oem_name.isnot(None))
     .group_by(Shipment.oem_name)
     .order_by(func.count(Shipment.id).desc()).all())


@reports_bp.route("/oem-performance")
@login_required
@staff_or_admin_required
def oem_performance():
    rows = _oem_perf_rows()
    return render_template("reports/oem_performance.html", rows=rows)


@reports_bp.route("/oem-performance.xlsx")
@login_required
@staff_or_admin_required
def oem_performance_xlsx():
    rows = _oem_perf_rows()
    headers = ["OEM", "Total", "Delivered", "Delayed",
               "Avg Readiness Delay (days)"]
    out = [[r[0], int(r[1]), int(r[2] or 0), int(r[3] or 0),
            round(float(r[4] or 0), 1)] for r in rows]
    buf = _wb_from_rows("OEM Performance", headers, out)
    return send_file(buf, as_attachment=True,
                     download_name="oem_performance.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# 4. Payment pending
def _payment_pending_rows():
    return (Shipment.query
            .filter(Shipment.payment_status.in_(["Pending", "Advance Paid"]))
            .order_by(Shipment.customer_po_date.desc()).all())


@reports_bp.route("/payment-pending")
@login_required
@staff_or_admin_required
def payment_pending():
    rows = _payment_pending_rows()
    return render_template("reports/payment_pending.html", rows=rows)


@reports_bp.route("/payment-pending.xlsx")
@login_required
@staff_or_admin_required
def payment_pending_xlsx():
    rows = _payment_pending_rows()
    headers = ["ID", "PO #", "OEM", "Bank", "PO Date",
               "Advance Date", "Balance Date", "Payment Status"]
    out = [[s.id, s.customer_po_number, s.oem_name, s.bank_name,
            s.customer_po_date, s.advance_payment_date,
            s.balance_payment_date, s.payment_status] for s in rows]
    buf = _wb_from_rows("Payment Pending", headers, out)
    return send_file(buf, as_attachment=True,
                     download_name="payment_pending.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# 5. Logistics delays
def _logistics_delay_rows():
    return (Shipment.query
            .filter((Shipment.sailing_delayed_days > 0) |
                    (Shipment.cc_delayed_days > 0))
            .order_by(Shipment.sailing_delayed_days.desc().nullslast()).all())


@reports_bp.route("/logistics-delays")
@login_required
@staff_or_admin_required
def logistics_delays():
    rows = _logistics_delay_rows()
    return render_template("reports/logistics_delays.html", rows=rows)


@reports_bp.route("/logistics-delays.xlsx")
@login_required
@staff_or_admin_required
def logistics_delays_xlsx():
    rows = _logistics_delay_rows()
    headers = ["ID", "PO #", "Logistics Partner", "Mode",
               "Committed Sailing", "Actual Sailing", "Sailing Delay",
               "Landing", "Clearance", "CC Delay"]
    out = [[s.id, s.customer_po_number, s.logistics_partner_name, s.mode,
            s.committed_sailing_date, s.actual_sailing_date,
            s.sailing_delayed_days,
            s.india_port_landing_date, s.custom_clearance_date,
            s.cc_delayed_days] for s in rows]
    buf = _wb_from_rows("Logistics Delays", headers, out)
    return send_file(buf, as_attachment=True,
                     download_name="logistics_delays.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# 6. Monthly MIS
def _monthly_rows(year):
    return (db.session.query(
        extract("month", Shipment.created_at).label("m"),
        func.count(Shipment.id).label("total"),
        func.sum(case((Shipment.status == "Delivered", 1), else_=0))
            .label("delivered"),
        func.sum(case((Shipment.delay_status == "Delayed", 1), else_=0))
            .label("delayed"),
        func.sum(func.coalesce(Shipment.shipment_quantity, 0)).label("qty"),
    ).filter(extract("year", Shipment.created_at) == year)
     .group_by("m").order_by("m").all())


@reports_bp.route("/monthly")
@login_required
@staff_or_admin_required
def monthly():
    year = request.args.get("year", date.today().year, type=int)
    rows = _monthly_rows(year)
    return render_template("reports/monthly.html", rows=rows, year=year)


@reports_bp.route("/monthly.xlsx")
@login_required
@staff_or_admin_required
def monthly_xlsx():
    year = request.args.get("year", date.today().year, type=int)
    rows = _monthly_rows(year)
    headers = ["Month", "Total", "Delivered", "Delayed", "Total Quantity"]
    out = [[int(r[0]), int(r[1]), int(r[2] or 0),
            int(r[3] or 0), int(r[4] or 0)] for r in rows]
    buf = _wb_from_rows(f"Monthly MIS {year}", headers, out)
    return send_file(buf, as_attachment=True,
                     download_name=f"monthly_mis_{year}.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
